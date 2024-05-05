import os
import time
import unittest

# Read the Excel file (replace with your filename and path)
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait

# Replace 'your_excel_file.xlsx' with the actual path and filename of your Excel file
workbook = load_workbook("Selenium Python\File Management\File Management.xlsx")

# Access a specific worksheet by name (assuming it's the first sheet)
worksheet = workbook["Sheet1"]

# Create an empty list to store the data (rows)
data_array = []

# Loop through rows in the worksheet (starting from row 2, assuming headers are in row 1)
for row in worksheet.iter_rows(min_row=2):
    # Create a dictionary to store the current row's data
    row_data = {}
    for cell in row:
        # Add cell value to the dictionary with column name as the key
        row_data[cell.column_letter] = cell.value
    # Append the row dictionary to the data array
    data_array.append(row_data)

for row in data_array:
    # Loop through each key-value pair in the current row dictionary
    for key, value in row.items():
        if value is None:
            row[key] = ""


class FileManagement(unittest.TestCase):

    def setUp(self):
        self.driver = webdriver.Chrome()
        driver = self.driver
        driver.get("https://school.moodledemo.net/login/index.php")
        driver.set_window_size(1024, 768)
        username_input = driver.find_element(By.ID, "username")
        username_input.send_keys("teacher")
        password_input = driver.find_element(By.ID, "password")
        password_input.send_keys("moodle")
        login_btn = driver.find_element(By.ID, "loginbtn")
        login_btn.click()

    def test_input_array(self):
        ################ Test Name Field ################

        name = data_array[0]["B"]
        driver = self.driver
        driver.get(
            "https://school.moodledemo.net/course/modedit.php?add=resource&type&course=59&section=0&return=0&beforemod=0"
        )
        name_input = driver.find_element(By.ID, "id_name")
        name_input.send_keys(name)
        driver.implicitly_wait(10)
        save_button = driver.find_element(By.ID, "id_submitbutton")
        save_button.click()
        driver.implicitly_wait(10)
        self.assertTrue(
            driver.find_element(By.ID, "id_error_name").text
            == "- You must supply a value here.",
        )

    def test_file_input(self):
        ################ Test Input ################
        test_case = data_array[4]

        name = test_case["B"]
        desc = test_case["C"]
        driver = self.driver
        driver.get(
            "https://school.moodledemo.net/course/modedit.php?add=resource&type&course=59&section=0&return=0&beforemod=0"
        )
        upload_file = os.path.abspath(
            os.path.join(os.path.dirname(__file__), "..", test_case["D"])
        )
        name_input = driver.find_element(By.ID, "id_name")
        name_input.send_keys(name)
        driver.implicitly_wait(10)
        driver.find_element(By.XPATH, "//div[4]/div/div[2]/div/div/i").click()

        driver.implicitly_wait(10)
        driver.find_element(By.LINK_TEXT, "Upload a file").click()
        driver.implicitly_wait(10)
        file_input = driver.find_element(By.CSS_SELECTOR, "input[type='file']")

        file_input.send_keys(upload_file)
        driver.find_element(
            By.XPATH, "//button[contains(.,'Upload this file')]"
        ).click()
        driver.implicitly_wait(10)
        save_button = driver.find_element(By.ID, "id_submitbutton")
        save_button.click()
        self.assertTrue(driver.find_element(By.XPATH, "//h1").text == name)

    def test_file_error(self, name="Test File"):
        ################ Test No File ################
        test_case = data_array[6]

        name = test_case["B"]
        desc = test_case["C"]
        driver = self.driver
        driver.get(
            "https://school.moodledemo.net/course/modedit.php?add=resource&type&course=59&section=0&return=0&beforemod=0"
        )
        upload_file = os.path.abspath(
            os.path.join(os.path.dirname(__file__), "..", test_case["D"])
        )
        name_input = driver.find_element(By.ID, "id_name")
        name_input.send_keys(name)
        driver.implicitly_wait(10)

        save_button = driver.find_element(By.ID, "id_submitbutton")
        save_button.click()
        self.assertTrue(
            driver.find_element(By.ID, "id_error_files").text == "Required",
        )

    def test_baseline(self):
        ################ Use-case Testing ################
        test_case = data_array[10]

        name = test_case["B"]
        desc = test_case["C"]
        driver = self.driver
        driver.get(
            "https://school.moodledemo.net/course/modedit.php?add=resource&type&course=59&section=0&return=0&beforemod=0"
        )
        upload_file = os.path.abspath(
            os.path.join(os.path.dirname(__file__), "..", test_case["D"])
        )
        name_input = driver.find_element(By.ID, "id_name")
        name_input.send_keys(name)
        driver.implicitly_wait(10)

        iframe = WebDriverWait(self.driver, 10).until(
            EC.visibility_of(driver.find_element(By.ID, "id_introeditor_ifr"))
        )

        # iframe = driver.find_element(By.ID, "id_introeditor_ifr")
        driver.switch_to.frame(iframe)
        desc_input = driver.find_element(By.XPATH, "//body[@id='tinymce']")
        desc_input.send_keys(desc)
        driver.implicitly_wait(10)
        driver.switch_to.default_content()

        driver.implicitly_wait(10)
        driver.find_element(By.XPATH, "//div[4]/div/div[2]/div/div/i").click()

        driver.implicitly_wait(10)
        driver.find_element(By.LINK_TEXT, "Upload a file").click()
        driver.implicitly_wait(10)
        file_input = driver.find_element(By.CSS_SELECTOR, "input[type='file']")

        file_input.send_keys(upload_file)
        driver.implicitly_wait(10)
        driver.find_element(
            By.XPATH, "//button[contains(.,'Upload this file')]"
        ).click()
        driver.implicitly_wait(10)
        save_button = driver.find_element(By.ID, "id_submitbutton")
        save_button.click()
        self.assertTrue(
            driver.find_element(By.XPATH, "//h1").text == name,
        )

    def test_basis_2(self):
        ################ Use-case Testing ################
        test_case = data_array[11]

        name = test_case["B"]
        desc = test_case["C"]
        driver = self.driver
        driver.get(
            "https://school.moodledemo.net/course/modedit.php?add=resource&type&course=59&section=0&return=0&beforemod=0"
        )
        upload_file = os.path.abspath(
            os.path.join(os.path.dirname(__file__), "..", test_case["D"])
        )
        name_input = driver.find_element(By.ID, "id_name")
        name_input.send_keys(name)
        iframe = WebDriverWait(self.driver, 10).until(
            EC.visibility_of(driver.find_element(By.ID, "id_introeditor_ifr"))
        )

        # iframe = driver.find_element(By.ID, "id_introeditor_ifr")
        driver.switch_to.frame(iframe)
        desc_input = driver.find_element(By.XPATH, "//body[@id='tinymce']")
        desc_input.send_keys(desc)
        driver.implicitly_wait(10)
        driver.switch_to.default_content()

        driver.implicitly_wait(10)
        driver.find_element(By.XPATH, "//div[4]/div/div[2]/div/div/i").click()

        driver.implicitly_wait(10)
        driver.find_element(By.LINK_TEXT, "Upload a file").click()
        file_input = driver.find_element(By.CSS_SELECTOR, "input[type='file']")
        driver.implicitly_wait(10)
        file_input.send_keys(upload_file)
        driver.find_element(
            By.XPATH, "//button[contains(.,'Upload this file')]"
        ).click()
        driver.implicitly_wait(10)
        save_button = driver.find_element(By.ID, "id_submitbutton2")
        save_button.click()
        self.assertTrue(
            len(driver.find_elements(By.XPATH, "//section[@id='region-main']/div"))
            != 0,
        )

    def test_basis_5(self):
        name = data_array[12][""]
        ################ Use-case Testing ################
        driver = self.driver
        driver.get(
            "https://school.moodledemo.net/course/modedit.php?add=resource&type&course=59&section=0&return=0&beforemod=0"
        )

        cancel_btn = driver.find_element(By.ID, "id_cancel")
        cancel_btn.click()
        self.assertTrue(
            len(driver.find_elements(By.XPATH, "//section[@id='region-main']/div"))
            != 0,
        )

    def tearDown(self):
        self.driver.close()


if __name__ == "__main__":
    unittest.main()
