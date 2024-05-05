import os
import time
import unittest

from openpyxl import load_workbook
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait

workbook = load_workbook("Selenium Python\Forum Post\Forum Post.xlsx")

# Access a specific worksheet by name (assuming it's the first sheet)
worksheet = workbook["Sheet1"]

# Create an empty list to store the data (rows)
data_array = []

# Loop through rows in the worksheet (starting from row 2, assuming headers are in row 1)
for row in worksheet.iter_rows(min_row=2):
    # Create a dictionary to store the current row's data
    row_data = {}
    for cell in row:
        # Add cell value to the dictionary with column name as the key
        row_data[cell.column_letter] = cell.value
    # Append the row dictionary to the data array
    data_array.append(row_data)

for row in data_array:
    # Loop through each key-value pair in the current row dictionary
    for key, value in row.items():
        if value is None:
            row[key] = ""


class ForumPost(unittest.TestCase):

    def setUp(self):
        self.driver = webdriver.Chrome()
        driver = self.driver
        driver.get("https://school.moodledemo.net/login/index.php")
        driver.set_window_size(1024, 768)
        username_input = driver.find_element(By.ID, "username")
        username_input.send_keys("student")
        password_input = driver.find_element(By.ID, "password")
        password_input.send_keys("moodle")
        login_btn = driver.find_element(By.ID, "loginbtn")
        login_btn.click()

    def test_input(self):
        name = data_array[0]["B"]
        msg = data_array[0]["C"]
        driver = self.driver
        driver.get("https://school.moodledemo.net/mod/forum/view.php?id=906")
        driver.find_element(By.LINK_TEXT, "Add discussion topic").click()
        driver.implicitly_wait(10)

        subject_input = driver.find_element(By.ID, "id_subject")
        subject_input.send_keys(name)
        driver.implicitly_wait(10)

        iframe = WebDriverWait(self.driver, 10).until(
            EC.visibility_of(driver.find_element(By.ID, "id_message_ifr"))
        )

        # iframe = driver.find_element(By.ID, "id_introeditor_ifr")
        driver.switch_to.frame(iframe)
        driver.implicitly_wait(10)
        msg_input = driver.find_element(By.XPATH, "//body[@id='tinymce']")
        msg_input.send_keys(msg)
        driver.implicitly_wait(10)
        driver.switch_to.default_content()

        driver.implicitly_wait(10)
        driver.find_element(By.ID, "id_submitbutton").click()
        if name == "":
            self.assertTrue(
                driver.find_element(By.ID, "id_error_subject").text == "Required",
            )
        elif msg == "":
            self.assertTrue(
                driver.find_element(By.ID, "id_error_message").text == "Required",
            )
        else:
            self.assertTrue(
                len(
                    driver.find_elements(
                        By.XPATH, "//p[contains(.,'Your post was successfully added.')]"
                    )
                )
                != 0,
            )

    def test_file_input(self):
        ################ Input Testing ################
        test_case = data_array[3]

        name = test_case["B"]
        msg = test_case["C"]

        driver = self.driver
        driver.get("https://school.moodledemo.net/mod/forum/view.php?id=906")

        driver.find_element(By.LINK_TEXT, "Add discussion topic").click()
        driver.implicitly_wait(10)

        subject_input = driver.find_element(By.ID, "id_subject")
        subject_input.send_keys(name)
        driver.implicitly_wait(10)

        iframe = WebDriverWait(self.driver, 10).until(
            EC.visibility_of(driver.find_element(By.ID, "id_message_ifr"))
        )
        driver.switch_to.frame(iframe)
        driver.implicitly_wait(10)
        msg_input = driver.find_element(By.XPATH, "//body[@id='tinymce']")
        msg_input.send_keys(msg)
        driver.implicitly_wait(10)
        driver.switch_to.default_content()

        driver.find_element(By.XPATH, "//button[@title='Image']").click()
        driver.implicitly_wait(10)
        driver.find_element(
            By.XPATH, "//button[contains(.,'Browse repositories')]"
        ).click()
        driver.implicitly_wait(10)

        file_input = driver.find_element(By.NAME, "repo_upload_file")
        upload_file = os.path.abspath(
            os.path.join(
                os.path.dirname(__file__),
                "..",
                test_case["D"],
            )
        )
        file_input.send_keys(upload_file)
        driver.implicitly_wait(10)
        driver.find_element(
            By.XPATH, "//button[contains(.,'Upload this file')]"
        ).click()
        error_modal = driver.find_element(
            By.XPATH, "//div[@class='moodle-exception-message']"
        )

        self.assertTrue(
            error_modal.text == "Video file (MP4) filetype cannot be accepted.",
        )

    def test_input_link(self):
        ################ Use-case Testing 1 ################
        test_case = data_array[6]

        name = test_case["B"]
        msg = test_case["C"]
        driver = self.driver
        driver.get("https://school.moodledemo.net/mod/forum/view.php?id=906")
        driver.find_element(By.LINK_TEXT, "Add discussion topic").click()
        driver.implicitly_wait(10)

        subject_input = driver.find_element(By.ID, "id_subject")
        subject_input.send_keys(name)
        driver.implicitly_wait(10)

        iframe = WebDriverWait(self.driver, 10).until(
            EC.visibility_of(driver.find_element(By.ID, "id_message_ifr"))
        )

        driver.switch_to.frame(iframe)
        driver.implicitly_wait(10)
        msg_input = driver.find_element(By.XPATH, "//body[@id='tinymce']")
        msg_input.send_keys(msg)
        driver.implicitly_wait(10)
        driver.switch_to.default_content()

        driver.find_element(By.XPATH, "//button[@title='Link']").click()
        url_text_input = driver.find_element(By.ID, "id_message_tiny_link_urltext")
        url_text_input.send_keys("this link")
        driver.implicitly_wait(10)
        url_entry_input = driver.find_element(By.ID, "id_message_tiny_link_urlentry")
        url_entry_input.send_keys("abcd.com")

        driver.find_element(By.XPATH, "button[contains(.,'Create link')]")

        driver.implicitly_wait(10)
        driver.find_element(By.ID, "id_submitbutton").click()
        self.assertTrue(
            len(
                driver.find_elements(
                    By.XPATH, "//p[contains(.,'Your post was successfully added.')]"
                )
            )
            != 0,
        )

    def test_baseline(self):
        ################ Use-case Testing 1 ################
        test_case = data_array[7]

        name = test_case["B"]
        msg = test_case["C"]
        driver = self.driver
        driver.get("https://school.moodledemo.net/mod/forum/view.php?id=906")
        driver.find_element(By.LINK_TEXT, "Add discussion topic").click()
        driver.implicitly_wait(10)

        subject_input = driver.find_element(By.ID, "id_subject")
        subject_input.send_keys(name)
        driver.implicitly_wait(10)

        iframe = WebDriverWait(self.driver, 10).until(
            EC.visibility_of(driver.find_element(By.ID, "id_message_ifr"))
        )

        # iframe = driver.find_element(By.ID, "id_introeditor_ifr")
        driver.switch_to.frame(iframe)
        driver.implicitly_wait(10)
        msg_input = driver.find_element(By.XPATH, "//body[@id='tinymce']")
        msg_input.send_keys(msg)
        driver.implicitly_wait(10)
        driver.switch_to.default_content()

        driver.implicitly_wait(10)
        driver.find_element(By.ID, "id_submitbutton").click()
        if name == "":
            self.assertTrue(
                driver.find_elements(By.ID, "id_error_subject").text == "Required",
            )
        elif msg == "":
            self.assertTrue(
                driver.find_elements(By.ID, "id_error_message").text == "Required",
            )
        else:
            self.assertTrue(
                len(
                    driver.find_elements(
                        By.XPATH, "//p[contains(.,'Your post was successfully added.')]"
                    )
                )
                != 0,
            )

    def test_basis_path_1(self):
        ################ Use-case Testing 2 ################
        test_case = data_array[8]

        name = test_case["B"]
        msg = test_case["C"]
        driver = self.driver
        driver.get("https://school.moodledemo.net/mod/forum/view.php?id=906")
        driver.find_element(By.LINK_TEXT, "Add discussion topic").click()
        driver.implicitly_wait(10)

        subject_input = driver.find_element(By.ID, "id_subject")
        subject_input.send_keys(name)
        driver.implicitly_wait(10)

        iframe = WebDriverWait(self.driver, 10).until(
            EC.visibility_of(driver.find_element(By.ID, "id_message_ifr"))
        )

        driver.switch_to.frame(iframe)
        driver.implicitly_wait(10)
        msg_input = driver.find_element(By.XPATH, "//body[@id='tinymce']")
        msg_input.send_keys(msg)
        driver.implicitly_wait(10)
        driver.switch_to.default_content()

        driver.find_element(By.ID, "id_advancedadddiscussion").click()

        input_select = driver.find_element(
            By.XPATH, "//input[contains(@placeholder, 'Enter tags...')]"
        )
        input_select.send_keys("test")
        input_select.send_keys(Keys.ENTER)

        driver.implicitly_wait(10)

        driver.find_element(By.ID, "id_submitbutton").click()
        self.assertTrue(
            len(
                driver.find_elements(
                    By.XPATH, "//p[contains(.,'Your post was successfully added.')]"
                )
            )
            != 0,
        )

    def test_basis_path_2(self):
        ################ Use-case Testing 3 ################
        test_case = data_array[9]

        name = test_case["B"]
        msg = test_case["C"]
        driver = self.driver
        driver.get("https://school.moodledemo.net/mod/forum/view.php?id=906")
        driver.find_element(By.LINK_TEXT, "Add discussion topic").click()
        driver.implicitly_wait(10)

        driver.find_element(By.ID, "id_cancelbtn").click()
        driver.implicitly_wait(10)
        try:
            driver.find_element(
                By.XPATH, "//div[@id='collapseAddForm' and contains(@class, 'show ')]"
            )
            # If the element is found, this code won't execute
            print("Element is present")
        except NoSuchElementException:
            print("Element is not present")

    def tearDown(self):
        self.driver.close()


if __name__ == "__main__":
    unittest.main()
